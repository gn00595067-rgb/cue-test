# cue_sheet_pro.py
# ============================================================
# Cue Sheet Pro - 擬真版（以 Excel 原生模板/COM 為主）
# 核心目標：最大化還原公司範例（字型/顏色/Logo/合併格/邊框）
#
# 依你最終規則：
# 1) Rate (Net) = 分區「總價」(不是單檔價)
#    Rate = int((List/Std)*Factor) * Spots
# 2) Spots 用 Net 算；未達標 penalty 會影響 Spots 計算
# 3) Package-cost 顯示用：
#    - 有選全省：顯示一格合併的打包價（全省 List 算；若未達標打包價 x1.1）
#      分區 Rate 不做 x1.1（避免價差太大讓客戶起疑）
#    - 沒選全省：逐列顯示；若未達標，Package-cost = round(Rate * 1.1)
# 4) Excel 產出：優先用 win32com 操作模板，保留 logo/shape
# 5) PDF 產出：用 Excel ExportAsFixedFormat，避免 HTML/LibreOffice 失真或出現 err
# ============================================================

import io
import os
import math
import tempfile
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional, Tuple

import streamlit as st

# --- Optional: openpyxl fallback（非 Windows/無 Excel 時） ---
try:
    import openpyxl
except Exception:
    openpyxl = None

# --- Windows Excel COM（最擬真） ---
HAS_EXCEL_COM = False
try:
    import pythoncom
    import win32com.client as win32
    HAS_EXCEL_COM = True
except Exception:
    HAS_EXCEL_COM = False


# ============================================================
# 1) Static DB（2026 新制）
# ============================================================

REGIONS_ORDER = ["北區", "桃竹苗", "中區", "雲嘉南", "高屏", "東區"]

REGION_DISPLAY = {
    "北區": "北區-北北基",
    "桃竹苗": "桃區-桃竹苗",
    "中區": "中區-中彰投",
    "雲嘉南": "雲嘉南區-雲嘉南",
    "高屏": "高屏區-高屏",
    "東區": "東區-宜花東",
    "全省": "全省",
}

STORE_COUNTS = {
    "全省": 4437,
    "北區": 1649,
    "桃竹苗": 779,
    "中區": 839,
    "雲嘉南": 499,
    "高屏": 490,
    "東區": 181,

    # 新鮮視（依你給的 key）
    "新鮮視_全省": 3124,
    "新鮮視_北區": 1127,
    "新鮮視_桃竹苗": 616,
    "新鮮視_中區": 528,
    "新鮮視_雲嘉南": 365,
    "新鮮視_高屏": 405,
    "新鮮視_東區": 83,

    # 家樂福
    "家樂福_量販": 68,
    "家樂福_超市": 249,
}

# Pricing DB：
# - List: 顯示用（價值定錨）
# - Net : 運算用（算 spots）
PRICING_DB = {
    "全家廣播": {
        "Std_Spots": 480,
        "Day_Part": "00:00-24:00",
        # region: [List, Net]
        "全省": [400000, 320000],
        "北區": [250000, 200000],
        "桃竹苗": [150000, 120000],
        "中區": [150000, 120000],
        "雲嘉南": [100000, 80000],
        "高屏": [100000, 80000],
        "東區": [62500, 50000],
    },
    "新鮮視": {
        "Std_Spots": 504,
        "Day_Part": "07:00-22:00",
        "全省": [150000, 120000],
        "北區": [150000, 120000],
        "桃竹苗": [120000, 96000],
        "中區": [90000, 72000],
        "雲嘉南": [75000, 60000],
        "高屏": [75000, 60000],
        "東區": [45000, 36000],
    },
    "家樂福": {
        "量販_全省": {"List": 300000, "Net": 250000, "Std_Spots": 420, "Day_Part": "09:00-23:00"},
        "超市_全省": {"List": 100000, "Net": 80000, "Std_Spots": 720, "Day_Part": "00:00-24:00"},
    }
}

SEC_FACTORS = {
    "全家廣播": {30: 1.0, 20: 0.85, 15: 0.65, 10: 0.5, 5: 0.25},
    "新鮮視":   {30: 3.0, 20: 2.0, 15: 1.5, 10: 1.0, 5: 0.5},
    "家樂福":   {30: 1.5, 20: 1.0, 15: 0.85, 10: 0.65, 5: 0.35},
}


# ============================================================
# 2) 基礎工具
# ============================================================

def ceil_div(a: float, b: float) -> int:
    if b == 0:
        return 0
    return int(math.ceil(a / b))

def evenize(n: int) -> int:
    if n <= 0:
        return 2
    return n if n % 2 == 0 else n + 1

def sec_factor(media: str, seconds: int) -> float:
    return SEC_FACTORS.get(media, {}).get(seconds, 1.0)

def calculate_schedule_even(total_spots: int, days: int) -> List[int]:
    """
    偶數排程（你的 v60.x 範例策略）
    - spots 強制偶數
    - 先除以2 分配，再乘回2
    """
    if days <= 0:
        return []
    total_spots = evenize(total_spots)
    half = total_spots // 2
    base = half // days
    rem = half % days
    arr = []
    for i in range(days):
        x = base + (1 if i < rem else 0)
        arr.append(x * 2)
    return arr

def station_display_name(media: str) -> str:
    # 依公司範例的 Station 換行寫法
    if media == "全家廣播":
        return "全家便利商店\n通路廣播廣告"
    if media == "新鮮視":
        return "全家便利商店\n新鮮視廣告"
    if media == "家樂福":
        return "家樂福"
    return media

def size_display(media: str, seconds: int) -> str:
    # 公司範例多用 "20秒"
    return f"{seconds}秒"

def program_display(media: str, region: str) -> str:
    if media == "新鮮視":
        key = f"新鮮視_{region}"
        v = STORE_COUNTS.get(key, 0)
        return f"{v:,}"
    if media == "全家廣播":
        v = STORE_COUNTS.get(region, 0)
        return f"{v:,}"
    if media == "家樂福":
        if region == "全省量販":
            return f"{STORE_COUNTS.get('家樂福_量販', 0):,}"
        if region == "全省超市":
            return f"{STORE_COUNTS.get('家樂福_超市', 0):,}"
    return "0"


# ============================================================
# 3) 核心運算（Spots）
# ============================================================

@dataclass
class RowOut:
    media: str
    seconds: int
    is_national: bool
    region: str
    location: str
    program: str
    daypart: str
    rate_total: Any     # int or str like "計量販"
    package_val: Any    # int or str, per-row (或合併時只有第一列有值)
    schedule: List[int]
    spots: int

@dataclass
class GroupOut:
    media: str
    seconds: int
    is_national: bool
    under_target: bool
    spots: int
    rows: List[RowOut]
    package_merged: bool
    package_merged_value: Optional[int]  # only if merged


def calc_spots_for_media(media: str, seconds: int, budget_alloc: float, regions_calc: List[str]) -> Tuple[int, bool]:
    """
    用 Net 算 spots：
    UnitCost(Net) = sum_over_regions (Net/Std)*Factor
    Spots_init = ceil(budget / UnitCost)
    under_target = Spots_init < Std
    Spots_final = ceil(budget / (UnitCost*(1.1 if under_target else 1)))
    Spots_final -> 偶數
    """
    factor = sec_factor(media, seconds)

    if media in ["全家廣播", "新鮮視"]:
        db = PRICING_DB[media]
        std = db["Std_Spots"]

        unit_sum = 0.0
        for r in regions_calc:
            net_price = db[r][1]
            unit_sum += (net_price / std) * factor

        spots_init = ceil_div(budget_alloc, unit_sum)
        under = spots_init < std
        mult = 1.1 if under else 1.0
        spots_final = ceil_div(budget_alloc, unit_sum * mult)
        spots_final = evenize(spots_final)
        return spots_final, under

    if media == "家樂福":
        # 用量販 Net 算 spots，超市用比例換算 spots
        db = PRICING_DB["家樂福"]
        base = db["量販_全省"]
        std = base["Std_Spots"]
        unit = (base["Net"] / std) * factor

        spots_init = ceil_div(budget_alloc, unit)
        under = spots_init < std
        mult = 1.1 if under else 1.0
        spots_final = ceil_div(budget_alloc, unit * mult)
        spots_final = evenize(spots_final)
        return spots_final, under

    return 0, False


def build_groups(
    total_budget_net: int,
    start_dt: date,
    end_dt: date,
    config_media: Dict[str, Dict[str, Any]],
) -> Tuple[List[GroupOut], Dict[str, Any]]:
    days = (end_dt - start_dt).days + 1
    if days <= 0:
        return [], {"days": 0}

    groups: List[GroupOut] = []
    debug = []

    # 依 share 分配預算
    for media, cfg in config_media.items():
        media_budget = total_budget_net * (cfg["share"] / 100.0)
        if media_budget <= 0:
            continue

        for sec, sec_share in cfg["sec_shares"].items():
            sec_budget = media_budget * (sec_share / 100.0)
            if sec_budget <= 0:
                continue

            if media in ["全家廣播", "新鮮視"]:
                is_nat = cfg["is_national"]
                # 計算用 regions：全省 -> ["全省"]；分區 -> 選到的那些
                regions_calc = ["全省"] if is_nat else cfg["regions"]
                # 顯示用 regions：全省 -> 展開 6 區；分區 -> 只顯示選到的
                regions_display = REGIONS_ORDER if is_nat else cfg["regions"]

                spots, under = calc_spots_for_media(media, sec, sec_budget, regions_calc)
                daypart = PRICING_DB[media]["Day_Part"]
                factor = sec_factor(media, sec)
                std = PRICING_DB[media]["Std_Spots"]

                schedule = calculate_schedule_even(spots, days)

                rows: List[RowOut] = []

                # Rate (Net) = 分區總價（用 List/Std*Factor 取 int 後乘 spots）
                # 注意：Rate 不做 x1.1（跟你附的 Cue_萬國通路 範例一致）
                for r in regions_display:
                    list_price = PRICING_DB[media][r][0] if not is_nat else PRICING_DB[media][r][0]
                    unit_list_int = int((list_price / std) * factor)
                    rate_total = unit_list_int * spots

                    # Package-cost：
                    # - 全省：要做合併格，per-row 先留空（只在第一列放 merged value）
                    # - 分區：逐列顯示；若未達標 package = round(rate * 1.1)
                    if is_nat:
                        pkg_val = None
                    else:
                        pkg_val = int(round(rate_total * (1.1 if under else 1.0)))

                    rows.append(RowOut(
                        media=media,
                        seconds=sec,
                        is_national=is_nat,
                        region=r,
                        location=REGION_DISPLAY.get(r, r),
                        program=program_display(media, r),
                        daypart=daypart,
                        rate_total=rate_total,
                        package_val=pkg_val,
                        schedule=schedule,
                        spots=spots
                    ))

                # 全省打包價（Package-cost 合併格）
                package_merged = is_nat
                package_merged_value = None
                if is_nat:
                    nat_list = PRICING_DB[media]["全省"][0]
                    unit_nat = int((nat_list / std) * factor)
                    base_pkg = unit_nat * spots
                    # 你的特例：全省時，未達標 -> 打包價要 x1.1；分區價不要 x1.1
                    package_merged_value = int(round(base_pkg * (1.1 if under else 1.0)))
                    # 把第一列 package_val 填入，後面靠 merge
                    if rows:
                        rows[0].package_val = package_merged_value

                groups.append(GroupOut(
                    media=media,
                    seconds=sec,
                    is_national=is_nat,
                    under_target=under,
                    spots=spots,
                    rows=rows,
                    package_merged=package_merged,
                    package_merged_value=package_merged_value
                ))

                debug.append({
                    "media": media, "sec": sec, "budget": sec_budget,
                    "spots": spots, "std": std, "under": under, "factor": factor,
                    "is_national": is_nat
                })

            elif media == "家樂福":
                # 家樂福：只有全省，但含量販/超市兩列
                spots_base, under = calc_spots_for_media("家樂福", sec, sec_budget, ["全省"])

                db = PRICING_DB["家樂福"]
                base = db["量販_全省"]
                sup = db["超市_全省"]
                factor = sec_factor("家樂福", sec)

                sch_base = calculate_schedule_even(spots_base, days)

                # 量販：Rate = int((List/Std)*Factor) * spots
                unit_list_int = int((base["List"] / base["Std_Spots"]) * factor)
                rate_total = unit_list_int * spots_base
                pkg_val = int(round(rate_total * (1.1 if under else 1.0)))

                row_base = RowOut(
                    media="家樂福",
                    seconds=sec,
                    is_national=True,
                    region="全省量販",
                    location="全省量販",
                    program=program_display("家樂福", "全省量販"),
                    daypart=base["Day_Part"],
                    rate_total=rate_total,
                    package_val=pkg_val,
                    schedule=sch_base,
                    spots=spots_base
                )

                # 超市：公司範例顯示「計量販」
                spots_sup = int(round(spots_base * (sup["Std_Spots"] / base["Std_Spots"])))
                spots_sup = evenize(spots_sup)
                sch_sup = calculate_schedule_even(spots_sup, days)

                row_sup = RowOut(
                    media="家樂福",
                    seconds=sec,
                    is_national=True,
                    region="全省超市",
                    location="全省超市",
                    program=program_display("家樂福", "全省超市"),
                    daypart=sup["Day_Part"],
                    rate_total="計量販",
                    package_val="計量販",
                    schedule=sch_sup,
                    spots=spots_sup
                )

                groups.append(GroupOut(
                    media="家樂福",
                    seconds=sec,
                    is_national=True,
                    under_target=under,
                    spots=spots_base,
                    rows=[row_base, row_sup],
                    package_merged=False,
                    package_merged_value=None
                ))

                debug.append({
                    "media": "家樂福", "sec": sec, "budget": sec_budget,
                    "spots": spots_base, "std": base["Std_Spots"], "under": under, "factor": factor,
                    "is_national": True
                })

    # 排序：全家廣播 -> 新鮮視 -> 家樂福；秒數小到大
    media_rank = {"全家廣播": 1, "新鮮視": 2, "家樂福": 3}
    groups.sort(key=lambda g: (media_rank.get(g.media, 99), g.seconds))

    meta = {
        "days": days,
        "debug": debug
    }
    return groups, meta


# ============================================================
# 4) Excel COM：用「公司範例模板」填值 + 匯出 PDF
# ============================================================

def excel_com_generate_from_template(
    template_path: str,
    client_name: str,
    products_str: str,
    start_dt: date,
    end_dt: date,
    groups: List[GroupOut],
    production_fee: int,
    budget_net: int
) -> Tuple[bytes, bytes]:
    """
    直接用 Excel COM 打開模板，填寫資料、維持格式、輸出 xlsx + pdf bytes
    """
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    tmp_dir = tempfile.mkdtemp(prefix="cue_sheet_")
    out_xlsx = os.path.join(tmp_dir, f"Cue_{client_name}.xlsx")
    out_pdf = os.path.join(tmp_dir, f"Cue_{client_name}.pdf")

    try:
        wb = excel.Workbooks.Open(os.path.abspath(template_path))
        ws = wb.Worksheets(1)  # 你給的公司範例是第一張

        # --- 1) 寫入 Header 區（依 Cue_萬國通路 範例格）---
        # A3 客戶名稱：, B3 value
        ws.Range("B3").Value = client_name
        ws.Range("B4").Value = products_str
        ws.Range("B5").Value = f"{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')}"
        # Medium（依 groups 用到的 media）
        used_media = []
        for g in groups:
            if g.media not in used_media:
                used_media.append(g.media)
        ws.Range("B6").Value = "、".join(used_media)

        # --- 2) 日期表頭（模板最多 31 天，從 H 欄開始）---
        days = (end_dt - start_dt).days + 1
        if days > 31:
            raise ValueError("此模板最多支援 31 天（公司範例格式）。若要跨月或>31天，需要你同意我重刻模板。")

        # 公司範例：第 8 列是「日期數字」、第 9 列是「週」
        day_num_row = 8
        weekday_row = 9
        day_start_col = 8  # H
        weekdays = ["一", "二", "三", "四", "五", "六", "日"]

        cur = start_dt
        for i in range(31):
            c = day_start_col + i
            if i < days:
                ws.Cells(day_num_row, c).Value = cur.day
                ws.Cells(weekday_row, c).Value = weekdays[cur.weekday()]
                cur += timedelta(days=1)
            else:
                ws.Cells(day_num_row, c).Value = ""
                ws.Cells(weekday_row, c).Value = ""

        # 檔次欄（AM=39）固定不動
        spots_col = 39

        # --- 3) 找到資料起始列/Total 列（依公司範例：Station header 在第 9 列）---
        header_row = 9
        data_start_row = 10

        # Total 字樣在 E 欄（公司範例）
        total_row = None
        used_rows = ws.UsedRange.Rows.Count
        # 從 data_start_row 往下找 "Total"
        for r in range(data_start_row, data_start_row + 400):
            if str(ws.Cells(r, 5).Value).strip() == "Total":
                total_row = r
                break
        if total_row is None:
            raise RuntimeError("找不到 Total 列（模板結構與公司範例不一致）。請改用公司範例 xlsx 當模板。")

        # --- 4) 計算需要的資料列數 ---
        flat_rows: List[RowOut] = []
        group_spans: List[Tuple[int, int, GroupOut]] = []  # (start_index, end_index, group)
        idx = 0
        for g in groups:
            start_i = idx
            for r in g.rows:
                flat_rows.append(r)
                idx += 1
            end_i = idx - 1
            group_spans.append((start_i, end_i, g))

        new_n = len(flat_rows)
        old_n = total_row - data_start_row

        # --- 5) 先把 data 區的 merge 解除（避免殘留合併造成錯位）---
        # 只解除資料區內的 merges（A~AM）
        def unmerge_in_rect(r1, c1, r2, c2):
            mr = ws.Range(ws.Cells(r1, c1), ws.Cells(r2, c2))
            # MergeCells 可能是 True/False；要取 Areas
            try:
                if mr.MergeCells:
                    mr.UnMerge()
            except Exception:
                pass

        # 比較安全做法：掃 usedrange 的 MergeCells.Areas
        try:
            mrange = ws.UsedRange.MergeCells
            if mrange:
                for area in mrange.Areas:
                    ar1 = area.Row
                    ac1 = area.Column
                    ar2 = ar1 + area.Rows.Count - 1
                    ac2 = ac1 + area.Columns.Count - 1
                    # 若 area 與資料區相交，解除合併
                    if not (ar2 < data_start_row or ar1 > total_row - 1 or ac2 < 1 or ac1 > spots_col):
                        area.UnMerge()
        except Exception:
            # 不致命：繼續
            pass

        # --- 6) 增減列數（在 Total 上方插入/刪除）並複製格式 ---
        # 用模板中 data_start_row 那一列當格式來源
        fmt_src_row = data_start_row

        if new_n > old_n:
            insert_cnt = new_n - old_n
            # 在 total_row 插入 insert_cnt 列
            ws.Rows(total_row).Resize(insert_cnt).Insert()
            # 把新插入列的格式複製成跟 fmt_src_row 一樣
            for k in range(insert_cnt):
                target_row = total_row + k
                ws.Rows(fmt_src_row).Copy()
                ws.Rows(target_row).PasteSpecial(Paste=-4122)  # xlPasteFormats = -4122
            excel.CutCopyMode = False
            total_row += insert_cnt

        elif new_n < old_n:
            delete_cnt = old_n - new_n
            ws.Rows(total_row - delete_cnt).Resize(delete_cnt).Delete()
            total_row -= delete_cnt

        # --- 7) 清空資料區內容（不破壞格式）---
        data_end_row = total_row - 1
        ws.Range(ws.Cells(data_start_row, 1), ws.Cells(data_end_row, spots_col)).ClearContents()

        # --- 8) 寫入資料列（A~G + 日程 H~AL + 檔次 AM）---
        day_start_col = 8  # H
        for i, r in enumerate(flat_rows):
            rr = data_start_row + i

            # A Station：由 group merge 寫入（先逐列寫，後面再合併）
            # B Location
            ws.Cells(rr, 2).Value = r.location
            # C Program
            ws.Cells(rr, 3).Value = r.program
            # D Day-part
            ws.Cells(rr, 4).Value = r.daypart
            # E Size
            ws.Cells(rr, 5).Value = size_display(r.media, r.seconds)
            # F Rate (Net) = 分區總價
            ws.Cells(rr, 6).Value = r.rate_total
            # G Package-cost (Net)
            ws.Cells(rr, 7).Value = r.package_val if r.package_val is not None else ""

            # H.. (31天) 日程
            for d in range(31):
                cc = day_start_col + d
                if d < len(r.schedule) and d < days:
                    ws.Cells(rr, cc).Value = r.schedule[d]
                else:
                    ws.Cells(rr, cc).Value = ""

            # AM 檔次
            ws.Cells(rr, spots_col).Value = r.spots

        # --- 9) 依 group 做 Station 合併 + Package-cost 合併 ---
        # Station (A) 合併：每個 group 一個 Station（公司範例行為）
        for (si, ei, g) in group_spans:
            r1 = data_start_row + si
            r2 = data_start_row + ei
            # Station (A)
            ws.Range(ws.Cells(r1, 1), ws.Cells(r2, 1)).Merge()
            ws.Cells(r1, 1).Value = station_display_name(g.media)

            # Package-cost (G) 合併：只有 multi-region 且 is_national 的 group
            if g.package_merged and g.media in ["全家廣播", "新鮮視"]:
                ws.Range(ws.Cells(r1, 7), ws.Cells(r2, 7)).Merge()
                ws.Cells(r1, 7).Value = g.package_merged_value if g.package_merged_value is not None else ""

        # --- 10) Total 列計算（依公司範例欄位）---
        # Total row：E欄有 "Total"，F=Rate 합計，G=Package 合計，H..AL=每日合計，AM=檔次合計
        total_rate = 0
        total_pkg = 0

        # Rate 合計：只加 int 的（"計量販" 不加）
        for r in flat_rows:
            if isinstance(r.rate_total, int):
                total_rate += r.rate_total

        # Package 合計：
        # - 全省 group（廣播/新鮮視）用 merged value
        # - 其他逐列加（int 才加）
        used_pkg_from_group = set()
        for g in groups:
            if g.package_merged and g.media in ["全家廣播", "新鮮視"]:
                if g.package_merged_value is not None:
                    total_pkg += g.package_merged_value
                used_pkg_from_group.add((g.media, g.seconds, g.is_national))
            else:
                for r in g.rows:
                    if isinstance(r.package_val, int):
                        total_pkg += r.package_val

        # 寫入 Total 列
        ws.Cells(total_row, 6).Value = total_rate
        ws.Cells(total_row, 7).Value = total_pkg

        # 每日 total：把每列 schedule 加總（只加顯示天數）
        day_totals = [0] * 31
        for r in flat_rows:
            for d in range(min(days, len(r.schedule), 31)):
                v = r.schedule[d]
                if isinstance(v, int):
                    day_totals[d] += v

        for d in range(31):
            ws.Cells(total_row, day_start_col + d).Value = day_totals[d] if d < days else ""

        # 檔次合計（公司範例是顯示在 AM）
        # 這裡做「每日總檔次」的總和（等同 total_row H..AL 加總的一半?）
        # 但公司範例通常填「總檔次」= sum(各 group spots) 或者你希望的總檔次
        # 依你貼的表格：Total 檔次是所有列檔次加總（含家樂福超市列也會加）
        ws.Cells(total_row, spots_col).Value = sum([r.spots for r in flat_rows if isinstance(r.spots, int)])

        # --- 11) Footer：製作 / Budget / VAT / Grand Total（位置依公司範例 G22:H25）---
        # 你若有不同模板，只要位置不一樣，這裡改座標即可
        vat = int(round((budget_net + production_fee) * 0.05))
        grand = budget_net + production_fee + vat

        ws.Range("H22").Value = production_fee
        ws.Range("H23").Value = budget_net
        ws.Range("H24").Value = vat
        ws.Range("H25").Value = grand

        # --- 12) 存檔 + 匯出 PDF ---
        wb.SaveAs(os.path.abspath(out_xlsx))
        ws.ExportAsFixedFormat(0, os.path.abspath(out_pdf))  # 0 = xlTypePDF

        wb.Close(SaveChanges=False)
        excel.Quit()

        with open(out_xlsx, "rb") as f:
            xlsx_bytes = f.read()
        with open(out_pdf, "rb") as f:
            pdf_bytes = f.read()

        return xlsx_bytes, pdf_bytes

    finally:
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


# ============================================================
# 5) openpyxl fallback（無 Excel 時，擬真度會差：shape/logo 可能消失）
# ============================================================

def openpyxl_generate_fallback(
    template_path: str,
    client_name: str,
    products_str: str,
    start_dt: date,
    end_dt: date,
    groups: List[GroupOut],
    production_fee: int,
    budget_net: int
) -> bytes:
    if openpyxl is None:
        raise RuntimeError("openpyxl 未安裝，且此環境也無 Excel COM。請安裝 openpyxl 或在 Windows 用 Excel COM。")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    # 盡量依公司範例填位
    ws["B3"].value = client_name
    ws["B4"].value = products_str
    ws["B5"].value = f"{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')}"
    used_media = []
    for g in groups:
        if g.media not in used_media:
            used_media.append(g.media)
    ws["B6"].value = "、".join(used_media)

    days = (end_dt - start_dt).days + 1
    if days > 31:
        raise ValueError("模板最多 31 天。")

    # 日期表頭
    weekdays = ["一", "二", "三", "四", "五", "六", "日"]
    cur = start_dt
    for i in range(31):
        col = openpyxl.utils.get_column_letter(8 + i)  # H start
        if i < days:
            ws[f"{col}8"].value = cur.day
            ws[f"{col}9"].value = weekdays[cur.weekday()]
            cur += timedelta(days=1)
        else:
            ws[f"{col}8"].value = ""
            ws[f"{col}9"].value = ""

    # 找 Total row
    total_row = None
    for r in range(10, 500):
        v = ws.cell(r, 5).value
        if isinstance(v, str) and v.strip() == "Total":
            total_row = r
            break
    if total_row is None:
        raise RuntimeError("找不到 Total row（fallback 無法處理此模板）。")

    data_start_row = 10
    spots_col = 39

    flat_rows: List[RowOut] = []
    group_spans: List[Tuple[int, int, GroupOut]] = []
    idx = 0
    for g in groups:
        si = idx
        for r in g.rows:
            flat_rows.append(r)
            idx += 1
        ei = idx - 1
        group_spans.append((si, ei, g))

    new_n = len(flat_rows)
    old_n = total_row - data_start_row

    # insert/delete rows
    if new_n > old_n:
        ws.insert_rows(total_row, amount=(new_n - old_n))
        total_row += (new_n - old_n)
    elif new_n < old_n:
        for _ in range(old_n - new_n):
            ws.delete_rows(total_row - 1, 1)
            total_row -= 1

    # 清空
    for r in range(data_start_row, total_row):
        for c in range(1, spots_col + 1):
            ws.cell(r, c).value = None

    # 寫入
    day_start_col = 8
    for i, r in enumerate(flat_rows):
        rr = data_start_row + i
        ws.cell(rr, 2).value = r.location
        ws.cell(rr, 3).value = r.program
        ws.cell(rr, 4).value = r.daypart
        ws.cell(rr, 5).value = size_display(r.media, r.seconds)
        ws.cell(rr, 6).value = r.rate_total
        ws.cell(rr, 7).value = r.package_val if r.package_val is not None else ""
        for d in range(31):
            cc = day_start_col + d
            if d < days and d < len(r.schedule):
                ws.cell(rr, cc).value = r.schedule[d]
            else:
                ws.cell(rr, cc).value = ""
        ws.cell(rr, spots_col).value = r.spots

    # merge
    for (si, ei, g) in group_spans:
        r1 = data_start_row + si
        r2 = data_start_row + ei
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        ws.cell(r1, 1).value = station_display_name(g.media)

        if g.package_merged and g.media in ["全家廣播", "新鮮視"]:
            ws.merge_cells(start_row=r1, start_column=7, end_row=r2, end_column=7)
            ws.cell(r1, 7).value = g.package_merged_value if g.package_merged_value is not None else ""

    # totals
    total_rate = sum([r.rate_total for r in flat_rows if isinstance(r.rate_total, int)])
    total_pkg = 0
    for g in groups:
        if g.package_merged and g.media in ["全家廣播", "新鮮視"]:
            if g.package_merged_value is not None:
                total_pkg += g.package_merged_value
        else:
            for r in g.rows:
                if isinstance(r.package_val, int):
                    total_pkg += r.package_val

    ws.cell(total_row, 6).value = total_rate
    ws.cell(total_row, 7).value = total_pkg

    # footer
    vat = int(round((budget_net + production_fee) * 0.05))
    grand = budget_net + production_fee + vat
    ws["H22"].value = production_fee
    ws["H23"].value = budget_net
    ws["H24"].value = vat
    ws["H25"].value = grand

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# 6) Streamlit UI
# ============================================================

st.set_page_config(layout="wide", page_title="Cue Sheet Pro（擬真版）")

st.title("📺 Cue Sheet Pro（擬真版：Excel 模板 COM 優先）")

with st.expander("⚙️ 重要說明（為什麼我這版會更像公司範例）", expanded=True):
    st.markdown(
        """
- **要做到字型/顏色/Logo/合併格完全像公司範例，必須用 Excel 原生模板 + COM 操作**（openpyxl 會掉 Shape/Logo，擬真度一定輸）。
- 這版會優先偵測 **Windows + Excel COM（pywin32）**，用同一份公司模板直接填值再匯出 PDF。
- 若你在雲端或沒裝 Excel，才會 fallback openpyxl（但那時擬真度會下降）。
        """
    )

col1, col2, col3 = st.columns([1.2, 1.2, 1])

with col1:
    client_name = st.text_input("客戶名稱", "萬國通路")
    start_date = st.date_input("開始日", date(2026, 1, 1))
with col2:
    end_date = st.date_input("結束日", date(2026, 1, 31))
    total_budget_net = st.number_input("總預算（未稅 Net）", min_value=0, value=1000000, step=10000)
with col3:
    production_fee = st.number_input("製作費", min_value=0, value=10000, step=1000)

days = (end_date - start_date).days + 1
st.info(f"📅 走期：{days} 天（模板最多 31 天）")

st.markdown("### 🧾 模板選擇（越像公司範例越重要）")
template_file = st.file_uploader(
    "請上傳「公司原版 Cue Excel（有 Logo、有配色、有合併格）」當模板（建議用你上傳的 Cue_萬國通路 (35).xlsx 那份作為基底）",
    type=["xlsx"]
)

default_template_path = None
# 如果你在本機同資料夾已經放公司範例模板，可以改這裡
# default_template_path = r"C:\...\Cue_萬國通路 (35).xlsx"

engine = st.selectbox(
    "輸出引擎（擬真度排序：Excel COM > openpyxl）",
    options=["AUTO（有 Excel 就用 COM）", "強制 Excel COM", "強制 openpyxl（不保證 logo/shape）"],
    index=0
)

st.markdown("### 2) 媒體投放設定（Share 加總需 = 100%）")

# UI：和你 v60.x 類似，但修正「未選全省不應出現全省/全部分區」
config_media: Dict[str, Dict[str, Any]] = {}

# 為了穩定：先讓使用者明確設定 share，避免「最後一個自動填滿」造成你難 debug
c1, c2, c3 = st.columns(3)

def sec_shares_ui(prefix: str, seconds_list: List[int]) -> Dict[int, int]:
    seconds_list = sorted(seconds_list)
    out = {}
    if not seconds_list:
        return out
    if len(seconds_list) == 1:
        out[seconds_list[0]] = 100
        return out

    left = 100
    for s in seconds_list[:-1]:
        v = st.slider(f"{s}秒佔比", 0, left, min(50, left), key=f"{prefix}_secshare_{s}")
        out[s] = v
        left -= v
    out[seconds_list[-1]] = left
    st.caption(f"🔹 {seconds_list[-1]}秒：{left}%")
    return out

with c1:
    st.subheader("📻 全家廣播")
    act = st.checkbox("啟用", value=True, key="fm_act")
    if act:
        is_nat = st.checkbox("全省聯播", value=True, key="fm_nat")
        regs = ["全省"] if is_nat else st.multiselect("區域", REGIONS_ORDER, default=["北區", "桃竹苗"], key="fm_regs")
        secs = st.multiselect("秒數", [5,10,15,20,30,35,45], default=[20], key="fm_secs")
        share = st.slider("預算佔比%", 0, 100, 70, key="fm_share")
        ss = sec_shares_ui("fm", secs)
        config_media["全家廣播"] = {"is_national": is_nat, "regions": regs if not is_nat else REGIONS_ORDER, "seconds": secs, "share": share, "sec_shares": ss}

with c2:
    st.subheader("📺 新鮮視")
    act = st.checkbox("啟用 ", value=True, key="fv_act")
    if act:
        is_nat = st.checkbox("全省聯播 ", value=False, key="fv_nat")
        regs = ["全省"] if is_nat else st.multiselect("區域 ", REGIONS_ORDER, default=["北區", "中區", "高屏"], key="fv_regs")
        secs = st.multiselect("秒數 ", [5,10,15,20,30,35,45], default=[10], key="fv_secs")
        share = st.slider("預算佔比% ", 0, 100, 20, key="fv_share")
        ss = sec_shares_ui("fv", secs)
        config_media["新鮮視"] = {"is_national": is_nat, "regions": regs if not is_nat else REGIONS_ORDER, "seconds": secs, "share": share, "sec_shares": ss}

with c3:
    st.subheader("🛒 家樂福")
    act = st.checkbox("啟用  ", value=True, key="cf_act")
    if act:
        secs = st.multiselect("秒數  ", [5,10,15,20,30,35,45], default=[20], key="cf_secs")
        share = st.slider("預算佔比%  ", 0, 100, 10, key="cf_share")
        ss = sec_shares_ui("cf", secs)
        config_media["家樂福"] = {"is_national": True, "regions": ["全省"], "seconds": secs, "share": share, "sec_shares": ss}

share_sum = sum([v["share"] for v in config_media.values()])
if share_sum != 100:
    st.warning(f"⚠️ 目前 share 加總 = {share_sum}%（需等於 100% 才能生成）")

# 產品字串（秒數集合）
secs_all = set()
for m, cfg in config_media.items():
    for s in cfg.get("seconds", []):
        secs_all.add(s)
products_str = "、".join([f"{s}秒" for s in sorted(secs_all)])

st.markdown("### 3) 生成結果")
colA, colB, colC = st.columns([1,1,1])
colA.metric("客戶預算（未稅）", f"{total_budget_net:,}")
vat = int(round((total_budget_net + production_fee) * 0.05))
colB.metric("VAT 5%", f"{vat:,}")
colC.metric("Grand Total（含稅）", f"{(total_budget_net+production_fee+vat):,}")

if st.button("🚀 生成 Cue（Excel + PDF）", type="primary", disabled=(share_sum != 100)):
    groups, meta = build_groups(total_budget_net, start_date, end_date, config_media)
    if not groups:
        st.error("沒有可輸出的資料（請檢查是否有選秒數、share、或預算為 0）。")
        st.stop()

    # 顯示 debug（避免你再被規則搞到崩潰）
    with st.expander("🧠 本次運算 Debug（檔次/是否未達標/全省 or 分區）", expanded=False):
        for d in meta["debug"]:
            st.write(
                f"{d['media']} {d['sec']}秒 | budget={d['budget']:.0f} | spots={d['spots']} | std={d['std']} | "
                f"{'未達標' if d['under'] else '達標'} | factor={d['factor']} | {'全省' if d['is_national'] else '分區'}"
            )

    # 決定模板來源
    if template_file is not None:
        tmp_template = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_template.write(template_file.read())
        tmp_template.close()
        template_path = tmp_template.name
    elif default_template_path and os.path.exists(default_template_path):
        template_path = default_template_path
    else:
        st.error("請上傳公司原版 Cue Excel 當模板（強烈建議用 Cue_萬國通路 的那份）。")
        st.stop()

    # 決定引擎
    use_com = False
    if engine == "AUTO（有 Excel 就用 COM）":
        use_com = HAS_EXCEL_COM
    elif engine == "強制 Excel COM":
        use_com = True
    else:
        use_com = False

    try:
        if use_com:
            if not HAS_EXCEL_COM:
                st.error("此環境無法使用 Excel COM（請確認 Windows + 已安裝 Excel + pip install pywin32）。")
                st.stop()

            xlsx_bytes, pdf_bytes = excel_com_generate_from_template(
                template_path=template_path,
                client_name=client_name,
                products_str=products_str,
                start_dt=start_date,
                end_dt=end_date,
                groups=groups,
                production_fee=production_fee,
                budget_net=total_budget_net
            )

            st.success("✅ 已以 Excel COM 生成（擬真度最高）")
            st.download_button(
                "📥 下載 Excel",
                data=xlsx_bytes,
                file_name=f"Cue_{client_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.download_button(
                "📄 下載 PDF（Excel 原生匯出，最像公司範例）",
                data=pdf_bytes,
                file_name=f"Cue_{client_name}.pdf",
                mime="application/pdf"
            )

        else:
            xlsx_bytes = openpyxl_generate_fallback(
                template_path=template_path,
                client_name=client_name,
                products_str=products_str,
                start_dt=start_date,
                end_dt=end_date,
                groups=groups,
                production_fee=production_fee,
                budget_net=total_budget_net
            )
            st.warning("⚠️ 已用 openpyxl fallback 生成（Logo/Shape/字型擬真度可能下降）。建議改用 Excel COM。")
            st.download_button(
                "📥 下載 Excel",
                data=xlsx_bytes,
                file_name=f"Cue_{client_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.exception(e)

    finally:
        try:
            if template_file is not None and os.path.exists(template_path):
                os.remove(template_path)
        except Exception:
            pass
